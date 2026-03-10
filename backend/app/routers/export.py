import io
import csv
import os
import zipfile
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from app.database import get_db
from app.models import User, UploadBatch, Image
from app.auth import get_current_user, require_admin, require_batch_or_admin

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/batch/{batch_id}/csv")
def export_batch_csv(
    batch_id: int,
    current_user: User = Depends(require_batch_or_admin),
    db: Session = Depends(get_db),
):
    """Export batch results as CSV. Requires batch or admin role."""
    batch, images = _get_batch_data(batch_id, current_user, db)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Image No.",
        "Original Filename",
        "AI Reading",
        "AI Confidence",
        "CV Reading",
        "CV Confidence",
        "Manual Correction",
        "Final Result",
        "Upload Time",
        "Patient Species",
        "Patient Age",
        "Patient Sex",
        "Patient Breed",
        "Patient Zip Code",
    ])

    for idx, img in enumerate(images, start=1):
        final = img.manual_correction or img.reading_result or img.cv_result or "Unclassified"
        pi = img.patient_info
        writer.writerow([
            idx,
            img.original_filename,
            img.reading_result or "",
            img.reading_confidence or "",
            img.cv_result or "",
            img.cv_confidence or "",
            img.manual_correction or "",
            final,
            img.created_at.strftime("%Y-%m-%d %H:%M:%S") if img.created_at else "",
            pi.species if pi and pi.species else "",
            pi.age if pi and pi.age else "",
            pi.sex if pi and pi.sex else "",
            pi.breed if pi and pi.breed else "",
            pi.zip_code if pi and pi.zip_code else "",
        ])

    output.seek(0)
    batch_label = batch.name or f"batch_{batch_id}"
    filename = f"{batch_label}_results.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/batch/{batch_id}/excel")
def export_batch_excel(
    batch_id: int,
    current_user: User = Depends(require_batch_or_admin),
    db: Session = Depends(get_db),
):
    """Export batch results as Excel. Requires batch or admin role."""
    batch, images = _get_batch_data(batch_id, current_user, db)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: Reading Results ---
    ws1 = wb.active
    ws1.title = "Reading Results"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    headers = [
        "Image No.",
        "Original Filename",
        "AI Reading",
        "AI Confidence",
        "CV Reading",
        "CV Confidence",
        "Manual Correction",
        "Final Result",
        "Upload Time",
        "Patient Species",
        "Patient Age",
        "Patient Sex",
        "Patient Breed",
        "Patient Zip Code",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, img in enumerate(images, start=1):
        final = img.manual_correction or img.reading_result or img.cv_result or "Unclassified"
        pi = img.patient_info
        row_data = [
            idx,
            img.original_filename,
            img.reading_result or "",
            img.reading_confidence or "",
            img.cv_result or "",
            img.cv_confidence or "",
            img.manual_correction or "",
            final,
            img.created_at.strftime("%Y-%m-%d %H:%M:%S") if img.created_at else "",
            pi.species if pi and pi.species else "",
            pi.age if pi and pi.age else "",
            pi.sex if pi and pi.sex else "",
            pi.breed if pi and pi.breed else "",
            pi.zip_code if pi and pi.zip_code else "",
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(images) + 2):
            cell_value = ws1.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws1.column_dimensions[chr(64 + col_idx)].width = min(max_length + 4, 40)

    # --- Sheet 2: Summary Statistics ---
    ws2 = wb.create_sheet(title="Summary")

    total = len(images)
    ai_count = sum(1 for img in images if img.reading_result)
    corrected_count = sum(1 for img in images if img.manual_correction)

    # Distribution
    final_dist = {}
    for img in images:
        final = img.manual_correction or img.reading_result or "Unclassified"
        final_dist[final] = final_dist.get(final, 0) + 1

    summary_data = [
        ["Batch Name", batch.name or "Untitled"],
        ["Total Images", total],
        ["AI Read", ai_count],
        ["Manually Corrected", corrected_count],
        ["Unclassified", total - max(ai_count, corrected_count, 0)],
        [],
        ["Category", "Count", "Percentage"],
    ]

    for category, count in sorted(final_dist.items(), key=lambda x: -x[1]):
        pct = f"{(count / total * 100):.1f}%" if total > 0 else "0%"
        summary_data.append([category, count, pct])

    for row_idx, row_data in enumerate(summary_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx <= 5 and col_idx == 1:
                cell.font = Font(bold=True)
            if row_idx == 7:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15

    # --- Sheet 3: Comparison (if applicable) ---
    both_images = [img for img in images if img.reading_result and img.manual_correction]
    if both_images:
        ws3 = wb.create_sheet(title="AI vs Manual")

        comp_headers = ["Image No.", "Filename", "AI Reading", "Manual Correction", "Match"]
        for col, header in enumerate(comp_headers, start=1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        match_count = 0
        for idx, img in enumerate(both_images, start=1):
            match = "Yes" if img.reading_result == img.manual_correction else "No"
            if match == "Yes":
                match_count += 1
            row_data = [idx, img.original_filename, img.reading_result, img.manual_correction, match]
            for col, value in enumerate(row_data, start=1):
                cell = ws3.cell(row=idx + 1, column=col, value=value)
                cell.alignment = Alignment(horizontal="center")
                if match == "No":
                    cell.font = Font(color="CC0000")

        # Accuracy row
        acc_row = len(both_images) + 3
        ws3.cell(row=acc_row, column=1, value="Accuracy").font = Font(bold=True)
        acc = f"{(match_count / len(both_images) * 100):.1f}%" if both_images else "N/A"
        ws3.cell(row=acc_row, column=2, value=acc)

        for col_idx in range(1, len(comp_headers) + 1):
            ws3.column_dimensions[chr(64 + col_idx)].width = 25

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    batch_label = batch.name or f"batch_{batch_id}"
    filename = f"{batch_label}_report.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/batch/{batch_id}/images")
def export_batch_images(
    batch_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Export preprocessed images as a ZIP archive. Admin only.

    Each image is named with its original filename for readability.
    If a preprocessed version exists, it is included; otherwise the
    original uploaded image is used as a fallback.
    """
    batch, images = _get_batch_data(batch_id, current_user, db)

    if not images:
        raise HTTPException(status_code=404, detail="No images in this batch")

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        used_names = {}
        for img in images:
            # Determine which file to include: preprocessed first, original as fallback
            if (
                img.is_preprocessed
                and img.preprocessed_path
                and os.path.exists(img.preprocessed_path)
            ):
                source_path = img.preprocessed_path
            elif os.path.exists(img.file_path):
                source_path = img.file_path
            else:
                continue

            # Use original filename in the ZIP; handle duplicates by appending a suffix
            base_name = img.original_filename
            if base_name in used_names:
                used_names[base_name] += 1
                name, ext = os.path.splitext(base_name)
                arc_name = f"{name}_{used_names[base_name]}{ext}"
            else:
                used_names[base_name] = 0
                arc_name = base_name

            zf.write(source_path, arc_name)

    buffer.seek(0)
    batch_label = batch.name or f"batch_{batch_id}"
    filename = f"{batch_label}_images.zip"

    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _get_batch_data(batch_id: int, current_user: User, db: Session):
    """Helper to get batch and images with ownership check. Admin can access any batch."""
    batch = (
        db.query(UploadBatch)
        .filter(UploadBatch.id == batch_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    if current_user.role != "admin" and batch.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    images = (
        db.query(Image)
        .options(joinedload(Image.patient_info))
        .filter(Image.batch_id == batch_id)
        .order_by(Image.id)
        .all()
    )
    return batch, images
